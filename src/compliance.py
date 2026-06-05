"""Generate compliance matrix Excel workbook from structured RFP analysis."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Colour fills ─────────────────────────────────────────────────────────────
_FILL_HEADER = PatternFill("solid", fgColor="1F3864")
_FILL_MANDATORY = PatternFill("solid", fgColor="FFE0E0")
_FILL_OPTIONAL = PatternFill("solid", fgColor="FFFFC0")
_FILL_RISK = PatternFill("solid", fgColor="FFD700")
_FILL_TBC = PatternFill("solid", fgColor="D3D3D3")
_FILL_SECTION = PatternFill("solid", fgColor="2E4057")
_FILL_HIGHLIGHT = PatternFill("solid", fgColor="FFEB9C")

# ── Fonts ─────────────────────────────────────────────────────────────────────
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
_FONT_BOLD = Font(bold=True)
_FONT_SECTION = Font(bold=True, color="FFFFFF", size=12)
_FONT_HIGHLIGHT = Font(bold=True, color="9C5700")

# ── Borders ───────────────────────────────────────────────────────────────────
_BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── Alignments ────────────────────────────────────────────────────────────────
_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
_ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Matrix sheet column definitions: (header label, column width) ─────────────
_MATRIX_COLUMNS: list[tuple[str, int]] = [
    ("Req ID", 10),
    ("Section", 18),
    ("Requirement", 60),
    ("Type", 12),
    ("Scoring Weight", 16),
    ("Compliance Status", 20),
    ("Nexus Solution", 28),
    ("Your Response", 45),
    ("Supporting Docs", 28),
    ("Remarks", 40),
]


def _bordered_wrap(cell: Any) -> None:
    """Apply thin border and word-wrap alignment to a single cell."""
    cell.border = _BORDER_THIN
    cell.alignment = _ALIGN_WRAP


def _write_matrix_sheet(
    ws: Any, requirements: list[dict[str, Any]]
) -> None:
    """Populate the Compliance Matrix worksheet.

    Args:
        ws: The openpyxl worksheet object to write into.
        requirements: List of requirement dicts from the RFP analysis.
    """
    # Header row
    for col_idx, (col_name, _) in enumerate(_MATRIX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.border = _BORDER_THIN
        cell.alignment = _ALIGN_HEADER

    # Data rows
    for row_idx, req in enumerate(requirements, start=2):
        req_type = str(req.get("type", "")).upper()
        risk_flag = bool(req.get("risk_flag", False))
        row_fill = _FILL_MANDATORY if req_type == "MANDATORY" else _FILL_OPTIONAL

        values = [
            req.get("id", ""),
            req.get("section", ""),
            req.get("requirement", ""),
            req_type,
            req.get("scoring_weight") or "",
            "TBC",
            req.get("nexus_solution") or "",
            "",
            "",
            req.get("remarks", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            _bordered_wrap(cell)

            if col_idx == 4:  # Type — bold, centred
                cell.font = _FONT_BOLD
                cell.alignment = _ALIGN_CENTER

            if col_idx == 6:  # Compliance Status — grey TBC
                cell.fill = _FILL_TBC
                cell.alignment = _ALIGN_CENTER

            if col_idx == 10 and risk_flag:  # Remarks — orange if risk
                cell.fill = _FILL_RISK

    # Column widths
    for col_idx, (_, width) in enumerate(_MATRIX_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row heights
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(requirements) + 2):
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "A2"


def _write_summary_sheet(ws: Any, analysis: dict[str, Any]) -> None:
    """Populate the RFP Summary worksheet.

    Args:
        ws: The openpyxl worksheet object to write into.
        analysis: Structured analysis dict from analyzer.analyze_rfp().
    """
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70

    current_row = 1

    def _section_header(title: str) -> None:
        nonlocal current_row
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.fill = _FILL_SECTION
        cell.font = _FONT_SECTION
        cell.border = _BORDER_THIN
        cell.alignment = Alignment(vertical="center")
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=2,
        )
        ws.row_dimensions[current_row].height = 24
        current_row += 1

    def _label_value(label: str, value: str) -> None:
        nonlocal current_row
        lc = ws.cell(row=current_row, column=1, value=label)
        lc.font = _FONT_BOLD
        lc.border = _BORDER_THIN
        lc.alignment = _ALIGN_WRAP
        vc = ws.cell(row=current_row, column=2, value=value)
        vc.border = _BORDER_THIN
        vc.alignment = _ALIGN_WRAP
        ws.row_dimensions[current_row].height = max(30, min(120, len(str(value)) // 2))
        current_row += 1

    def _bullet_list(items: list[str]) -> None:
        nonlocal current_row
        for item in items:
            cell = ws.cell(row=current_row, column=2, value=f"• {item}")
            cell.border = _BORDER_THIN
            cell.alignment = _ALIGN_WRAP
            ws.row_dimensions[current_row].height = 20
            current_row += 1

    def _highlighted_kv(label: str, value: str, highlight: bool) -> None:
        nonlocal current_row
        lc = ws.cell(row=current_row, column=1, value=label)
        lc.font = _FONT_BOLD
        lc.border = _BORDER_THIN
        lc.alignment = _ALIGN_WRAP
        vc = ws.cell(row=current_row, column=2, value=value)
        vc.border = _BORDER_THIN
        if highlight:
            vc.fill = _FILL_HIGHLIGHT
            vc.font = _FONT_HIGHLIGHT
        current_row += 1

    # Executive Summary
    _section_header("EXECUTIVE SUMMARY")
    _label_value("Summary", analysis.get("executive_summary", ""))
    current_row += 1

    # Key Details
    _section_header("KEY DETAILS")
    deadline = analysis.get("submission_deadline") or "Not specified"
    budget = analysis.get("budget") or "Not specified"
    _highlighted_kv("Submission Deadline", deadline, deadline != "Not specified")
    _highlighted_kv("Budget", budget, budget != "Not specified")
    _label_value("Industry Vertical", analysis.get("industry_vertical", ""))
    current_row += 1

    # Evaluation Criteria
    _section_header("EVALUATION CRITERIA")
    criteria = analysis.get("evaluation_criteria", [])
    if criteria:
        hdr_c = ws.cell(row=current_row, column=1, value="Criterion")
        hdr_c.font = _FONT_BOLD
        hdr_c.border = _BORDER_THIN
        hdr_w = ws.cell(row=current_row, column=2, value="Weight")
        hdr_w.font = _FONT_BOLD
        hdr_w.border = _BORDER_THIN
        current_row += 1
        for item in criteria:
            cc = ws.cell(row=current_row, column=1, value=item.get("criterion", ""))
            cc.border = _BORDER_THIN
            cc.alignment = _ALIGN_WRAP
            wc = ws.cell(row=current_row, column=2, value=item.get("weight", ""))
            wc.border = _BORDER_THIN
            current_row += 1
    else:
        nc = ws.cell(row=current_row, column=2, value="Not specified")
        nc.border = _BORDER_THIN
        current_row += 1
    current_row += 1

    # Customer Objectives
    _section_header("CUSTOMER OBJECTIVES")
    objectives = analysis.get("customer_objectives", [])
    _bullet_list(objectives if objectives else ["Not specified"])
    current_row += 1

    # Win Themes
    _section_header("WIN THEMES")
    _bullet_list(analysis.get("win_themes", []) or ["Not identified"])
    current_row += 1

    # Risk Areas
    _section_header("RISK AREAS")
    _bullet_list(analysis.get("risk_areas", []) or ["None identified"])


def generate_excel(analysis: dict[str, Any], output_dir: str) -> str:
    """Generate a compliance matrix Excel workbook from structured RFP analysis.

    Creates two sheets:
      Sheet 1 — Compliance Matrix: one row per requirement with status tracking.
      Sheet 2 — RFP Summary: executive summary, criteria, win themes, risk areas.

    Filename pattern: {rfp_stem_20chars}_Compliance_{YYYYMMDD}.xlsx

    Args:
        analysis: Structured dict from analyzer.analyze_rfp(), expected to include
            a '_rfp_filename' key set by the caller.
        output_dir: Directory where the .xlsx file is saved. Created if absent.

    Returns:
        Absolute path to the saved Excel file.

    Raises:
        OSError: If the output directory cannot be created or the file cannot be written.
    """
    out_path = Path(output_dir)
    try:
        out_path.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        logger.error("Cannot create output directory '%s': %s", output_dir, exc)
        raise

    rfp_name = analysis.get("_rfp_filename", "rfp")
    stem = Path(rfp_name).stem[:20]
    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"{stem}_Compliance_{date_str}.xlsx"
    file_path = out_path / filename

    logger.info("Generating compliance Excel: %s", file_path)

    wb = Workbook()

    ws_matrix = wb.active
    ws_matrix.title = "Compliance Matrix"
    requirements = analysis.get("requirements", [])
    _write_matrix_sheet(ws_matrix, requirements)
    logger.info("Compliance Matrix: %d requirements written", len(requirements))

    ws_summary = wb.create_sheet("RFP Summary")
    _write_summary_sheet(ws_summary, analysis)
    logger.info("RFP Summary sheet written")

    try:
        wb.save(str(file_path))
    except OSError as exc:
        logger.error("Failed to save Excel '%s': %s", file_path, exc)
        raise

    resolved = str(file_path.resolve())
    logger.info("Excel saved: %s", resolved)
    return resolved
